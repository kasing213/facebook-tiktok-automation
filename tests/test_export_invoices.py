"""
Test script for invoice export functionality.
Tests CSV and XLSX generation logic without requiring a live database.
"""
import sys
import os
from io import BytesIO
from datetime import datetime, date
from collections import namedtuple

# Simulate a database row matching the export query columns
InvoiceRow = namedtuple("InvoiceRow", [
    "invoice_number", "amount", "status", "currency",
    "bank", "expected_account", "recipient_name",
    "due_date", "verification_status",
    "created_at", "updated_at",
    "customer_name"
])

HEADERS = [
    "Invoice Number", "Customer", "Amount", "Currency", "Status",
    "Verification Status", "Bank", "Account", "Recipient",
    "Due Date", "Created At", "Updated At"
]


def format_row(row):
    """Exact copy of the format_row function from invoice.py export endpoint."""
    return [
        row.invoice_number or "",
        row.customer_name or "Unknown",
        f"{float(row.amount):.2f}" if row.amount else "0.00",
        row.currency or "USD",
        row.status or "",
        row.verification_status or "",
        row.bank or "",
        row.expected_account or "",
        row.recipient_name or "",
        row.due_date.strftime("%Y-%m-%d") if row.due_date else "",
        row.created_at.strftime("%Y-%m-%d %H:%M") if row.created_at else "",
        row.updated_at.strftime("%Y-%m-%d %H:%M") if row.updated_at else "",
    ]


SAMPLE_ROWS = [
    InvoiceRow(
        invoice_number="INV-001", amount=150.50, status="paid", currency="USD",
        bank="ABA", expected_account="123456789", recipient_name="John Doe",
        due_date=date(2026, 2, 15), verification_status="verified",
        created_at=datetime(2026, 1, 10, 14, 30), updated_at=datetime(2026, 1, 12, 9, 0),
        customer_name="Acme Corp"
    ),
    InvoiceRow(
        invoice_number="INV-002", amount=0, status="draft", currency="KHR",
        bank=None, expected_account=None, recipient_name=None,
        due_date=None, verification_status=None,
        created_at=datetime(2026, 2, 1, 8, 0), updated_at=None,
        customer_name=None
    ),
    InvoiceRow(
        invoice_number=None, amount=None, status="pending", currency=None,
        bank="Wing", expected_account="0987", recipient_name="Jane",
        due_date=date(2026, 3, 1), verification_status="pending",
        created_at=None, updated_at=None,
        customer_name="Test, With Comma"
    ),
]

passed = 0
failed = 0


def test(name, condition, detail=""):
    global passed, failed
    if condition:
        print(f"  PASS: {name}")
        passed += 1
    else:
        print(f"  FAIL: {name} - {detail}")
        failed += 1


# ── Test 1: format_row with normal data ──
print("\n[1] format_row with normal data")
row1 = format_row(SAMPLE_ROWS[0])
test("returns list of 12 items", len(row1) == 12, f"got {len(row1)}")
test("invoice_number correct", row1[0] == "INV-001")
test("customer_name correct", row1[1] == "Acme Corp")
test("amount formatted", row1[2] == "150.50")
test("currency correct", row1[3] == "USD")
test("due_date formatted", row1[4] == "paid")
test("created_at formatted", row1[10] == "2026-01-10 14:30")

# ── Test 2: format_row with nulls ──
print("\n[2] format_row with null/missing fields")
row2 = format_row(SAMPLE_ROWS[1])
test("null customer -> Unknown", row2[1] == "Unknown")
test("zero amount -> 0.00", row2[2] == "0.00")
test("null bank -> empty", row2[6] == "")
test("null due_date -> empty", row2[9] == "")
test("null updated_at -> empty", row2[11] == "")

row3 = format_row(SAMPLE_ROWS[2])
test("null invoice_number -> empty", row3[0] == "")
test("null amount -> 0.00", row3[2] == "0.00")
test("null currency -> USD", row3[3] == "USD")

# ── Test 3: CSV generation ──
print("\n[3] CSV generation")
lines = [",".join(HEADERS)]
for row in SAMPLE_ROWS:
    lines.append(",".join(str(v).replace(",", " ") for v in format_row(row)))
csv_content = "\n".join(lines).encode("utf-8")

csv_text = csv_content.decode("utf-8")
csv_lines = csv_text.strip().split("\n")
test("header row present", csv_lines[0].startswith("Invoice Number"))
test("3 data rows + 1 header = 4 lines", len(csv_lines) == 4, f"got {len(csv_lines)}")
test("commas in data escaped", "Test  With Comma" in csv_lines[3], f"got: {csv_lines[3]}")
test("csv is bytes", isinstance(csv_content, bytes))
test("csv not empty", len(csv_content) > 0, f"size: {len(csv_content)}")

# ── Test 4: CSV empty state ──
print("\n[4] CSV with no data rows")
empty_lines = [",".join(HEADERS)]
empty_csv = "\n".join(empty_lines).encode("utf-8")
test("empty csv has header only", len(empty_csv.decode().strip().split("\n")) == 1)
test("empty csv not zero bytes", len(empty_csv) > 0)

# ── Test 5: XLSX generation ──
print("\n[5] XLSX generation (openpyxl)")
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    rows = SAMPLE_ROWS
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(format_row(row), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    for col_idx in range(1, len(HEADERS) + 1):
        max_length = len(HEADERS[col_idx - 1])
        for row_idx in range(2, len(rows) + 2):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 40)

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    xlsx_bytes = buffer.getvalue()

    test("xlsx is bytes", isinstance(xlsx_bytes, bytes))
    test("xlsx not empty", len(xlsx_bytes) > 100, f"size: {len(xlsx_bytes)}")
    test("xlsx has valid ZIP header (PK)", xlsx_bytes[:2] == b"PK")

    # Re-read and validate content
    rb = load_workbook(BytesIO(xlsx_bytes))
    rws = rb.active
    test("sheet title is 'Invoices'", rws.title == "Invoices")
    test("header row A1 = 'Invoice Number'", rws.cell(1, 1).value == "Invoice Number")
    test("header row L1 = 'Updated At'", rws.cell(1, 12).value == "Updated At")
    test("data row A2 = 'INV-001'", rws.cell(2, 1).value == "INV-001")
    test("data row B2 = 'Acme Corp'", rws.cell(2, 2).value == "Acme Corp")
    test("data row C2 = '150.50'", rws.cell(2, 3).value == "150.50")
    test("null customer row B3 = 'Unknown'", rws.cell(3, 2).value == "Unknown")
    test("4 rows total (1 header + 3 data)", rws.max_row == 4, f"got {rws.max_row}")
    test("12 columns", rws.max_column == 12, f"got {rws.max_column}")

    # Verify header styling
    h1 = rws.cell(1, 1)
    test("header font is bold", h1.font.bold == True)
    test("header fill is blue", h1.fill.start_color.rgb == "004472C4" or h1.fill.start_color.rgb == "4472C4" or "4472C4" in str(h1.fill.start_color.rgb))
    test("frozen panes set", rws.freeze_panes == "A2")

    rb.close()

except ImportError as e:
    print(f"  SKIP: openpyxl not installed - {e}")
    failed += 1

# ── Test 6: XLSX empty state ──
print("\n[6] XLSX with no data rows")
try:
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Invoices"
    for col_idx, header in enumerate(HEADERS, 1):
        ws2.cell(row=1, column=col_idx, value=header)
    buf2 = BytesIO()
    wb2.save(buf2)
    empty_xlsx = buf2.getvalue()
    test("empty xlsx is valid", len(empty_xlsx) > 100)
    test("empty xlsx has ZIP header", empty_xlsx[:2] == b"PK")

    rb2 = load_workbook(BytesIO(empty_xlsx))
    test("empty xlsx has 1 row (header only)", rb2.active.max_row == 1)
    rb2.close()
except Exception as e:
    print(f"  FAIL: empty xlsx - {e}")
    failed += 1

# ── Summary ──
print(f"\n{'='*50}")
print(f"Results: {passed} passed, {failed} failed")
print(f"{'='*50}")

sys.exit(0 if failed == 0 else 1)

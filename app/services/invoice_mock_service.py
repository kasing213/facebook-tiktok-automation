# app/services/invoice_mock_service.py
"""
Mock Invoice API service for testing.

Provides in-memory CRUD operations that simulate the external Invoice API.
Data is stored in memory and lost on restart - suitable for development/testing.
"""

from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
from uuid import uuid4


# Tenant-aware in-memory data stores
# Structure: {tenant_id: {entity_id: entity_data}}
_customers: Dict[str, Dict[str, Dict[str, Any]]] = {}
_invoices: Dict[str, Dict[str, Dict[str, Any]]] = {}
_invoice_counters: Dict[str, int] = {}  # tenant_id -> counter


def _generate_id() -> str:
    """Generate a unique ID."""
    return str(uuid4())


def _now_iso() -> str:
    """Get current timestamp in ISO format."""
    return datetime.utcnow().isoformat() + "Z"


def _ensure_tenant_data(tenant_id: str) -> None:
    """Ensure tenant has initialized data structures."""
    tenant_str = str(tenant_id)
    if tenant_str not in _customers:
        _customers[tenant_str] = {}
    if tenant_str not in _invoices:
        _invoices[tenant_str] = {}
    if tenant_str not in _invoice_counters:
        _invoice_counters[tenant_str] = 1000


# ============================================================================
# Customer Operations
# ============================================================================

def list_customers(
    tenant_id: str,
    limit: int = 50,
    skip: int = 0,
    search: Optional[str] = None
) -> List[Dict]:
    """List customers with optional search and pagination (tenant-isolated)."""
    _ensure_tenant_data(tenant_id)
    customers = list(_customers[str(tenant_id)].values())

    if search:
        search_lower = search.lower()
        customers = [
            c for c in customers
            if search_lower in c["name"].lower()
            or (c.get("email") and search_lower in c["email"].lower())
            or (c.get("company") and search_lower in c["company"].lower())
        ]

    # Sort by created_at descending
    customers.sort(key=lambda x: x.get("created_at", ""), reverse=True)

    return customers[skip:skip + limit]


def get_customer(tenant_id: str, customer_id: str) -> Optional[Dict]:
    """Get a single customer by ID (tenant-isolated)."""
    _ensure_tenant_data(tenant_id)
    return _customers[str(tenant_id)].get(customer_id)


def create_customer(tenant_id: str, data: Dict) -> Dict:
    """Create a new customer (tenant-isolated)."""
    _ensure_tenant_data(tenant_id)
    customer_id = _generate_id()
    customer = {
        "id": customer_id,
        "tenant_id": str(tenant_id),
        "name": data["name"],
        "email": data.get("email"),
        "phone": data.get("phone"),
        "address": data.get("address"),
        "company": data.get("company"),
        "notes": data.get("notes"),
        "created_at": _now_iso(),
        "updated_at": _now_iso()
    }
    _customers[str(tenant_id)][customer_id] = customer
    return customer


def update_customer(tenant_id: str, customer_id: str, data: Dict) -> Optional[Dict]:
    """Update an existing customer (tenant-isolated)."""
    _ensure_tenant_data(tenant_id)
    tenant_customers = _customers[str(tenant_id)]
    if customer_id not in tenant_customers:
        return None

    customer = tenant_customers[customer_id]
    for key, value in data.items():
        if value is not None:
            customer[key] = value
    customer["updated_at"] = _now_iso()

    return customer


def delete_customer(tenant_id: str, customer_id: str) -> bool:
    """Delete a customer (tenant-isolated)."""
    _ensure_tenant_data(tenant_id)
    tenant_customers = _customers[str(tenant_id)]
    if customer_id in tenant_customers:
        del tenant_customers[customer_id]
        return True
    return False


# ============================================================================
# Invoice Operations
# ============================================================================

def list_invoices(
    tenant_id: str,
    limit: int = 50,
    skip: int = 0,
    customer_id: Optional[str] = None,
    status: Optional[str] = None
) -> List[Dict]:
    """List invoices with optional filtering and pagination (tenant-isolated)."""
    _ensure_tenant_data(tenant_id)
    invoices = list(_invoices[str(tenant_id)].values())

    if customer_id:
        invoices = [i for i in invoices if i["customer_id"] == customer_id]
    if status:
        invoices = [i for i in invoices if i["status"] == status]

    # Add customer info to each invoice (from same tenant)
    tenant_customers = _customers[str(tenant_id)]
    for inv in invoices:
        inv["customer"] = tenant_customers.get(inv["customer_id"])

    # Sort by created_at descending
    invoices.sort(key=lambda x: x.get("created_at", ""), reverse=True)

    return invoices[skip:skip + limit]


def get_invoice(tenant_id: str, invoice_id: str) -> Optional[Dict]:
    """Get a single invoice by ID with customer info (tenant-isolated)."""
    _ensure_tenant_data(tenant_id)
    # First check in-memory mock data
    tenant_invoices = _invoices[str(tenant_id)]
    invoice = tenant_invoices.get(invoice_id)
    if invoice:
        # Return a copy with customer info (from same tenant)
        invoice = invoice.copy()
        tenant_customers = _customers[str(tenant_id)]
        invoice["customer"] = tenant_customers.get(invoice["customer_id"])
        return invoice

    # Fallback to PostgreSQL database if not found in mock data
    try:
        from app.core.db import get_db_sync
        from sqlalchemy import text
        import json

        db = next(get_db_sync())

        # Query invoice with customer data (matching actual database schema)
        # CRITICAL: Filter by tenant_id to prevent cross-tenant data access
        result = db.execute(text("""
            SELECT
                i.id,
                i.invoice_number,
                i.customer_id,
                i.status,
                i.amount,
                i.currency,
                i.items,
                i.due_date,
                i.created_at,
                i.bank,
                i.expected_account,
                i.recipient_name,
                i.verification_status,
                c.name as customer_name,
                c.email as customer_email,
                c.phone as customer_phone,
                c.address as customer_address
            FROM invoice.invoice i
            LEFT JOIN invoice.customer c ON i.customer_id = c.id
            WHERE i.id = :invoice_id AND i.tenant_id = :tenant_id
        """), {"invoice_id": invoice_id, "tenant_id": str(tenant_id)}).fetchone()

        if result:
            # Parse items JSON (handle both string and already-parsed list)
            if result.items:
                if isinstance(result.items, str):
                    items = json.loads(result.items)
                else:
                    # SQLAlchemy already deserialized JSONB to list
                    items = result.items
            else:
                items = []

            # Calculate totals from items if needed
            subtotal = sum(
                item.get("quantity", 0) * item.get("unit_price", 0)
                for item in items
            )
            total = float(result.amount) if result.amount else subtotal

            # Build invoice dict
            invoice = {
                "id": str(result.id),
                "invoice_number": result.invoice_number,
                "customer_id": str(result.customer_id) if result.customer_id else None,
                "status": result.status,
                "amount": total,
                "total": total,
                "subtotal": subtotal,
                "discount": 0,
                "currency": result.currency or "KHR",
                "items": items,
                "notes": None,
                "due_date": result.due_date.isoformat() if result.due_date else None,
                "invoice_date": result.created_at.isoformat() if result.created_at else None,
                "created_at": result.created_at.isoformat() if result.created_at else None,
                "updated_at": result.created_at.isoformat() if result.created_at else None,
                "bank": result.bank,
                "expected_account": result.expected_account,
                "recipient_name": result.recipient_name,
                "verification_status": result.verification_status,
                "customer": {
                    "id": str(result.customer_id) if result.customer_id else None,
                    "name": result.customer_name,
                    "email": result.customer_email,
                    "phone": result.customer_phone,
                    "address": result.customer_address
                } if result.customer_name else None
            }
            return invoice

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Failed to query invoice {invoice_id} from database: {e}")

    return None


def create_invoice(tenant_id: str, data: Dict) -> Dict:
    """Create a new invoice with auto-calculated totals (tenant-isolated)."""
    _ensure_tenant_data(tenant_id)
    tenant_str = str(tenant_id)

    invoice_id = _generate_id()
    _invoice_counters[tenant_str] += 1

    # Process items
    items = data.get("items", [])
    processed_items = []

    for item in items:
        quantity = float(item.get("quantity", 1))
        unit_price = float(item.get("unit_price", 0))
        tax_rate = float(item.get("tax_rate", 0))
        item_total = quantity * unit_price

        processed_items.append({
            "id": _generate_id(),
            "description": item.get("description", ""),
            "quantity": quantity,
            "unit_price": unit_price,
            "tax_rate": tax_rate,
            "total": round(item_total, 2)
        })

    # Calculate totals
    subtotal = sum(item["total"] for item in processed_items)
    tax_amount = sum(
        item["total"] * (item.get("tax_rate", 0) / 100)
        for item in processed_items
    )
    discount = float(data.get("discount", 0))
    discount_amount = subtotal * (discount / 100) if discount > 0 else 0
    total = subtotal + tax_amount - discount_amount

    invoice = {
        "id": invoice_id,
        "invoice_number": f"INV-{_invoice_counter:05d}",
        "customer_id": data["customer_id"],
        "items": processed_items,
        "subtotal": round(subtotal, 2),
        "tax_rate": sum(item.get("tax_rate", 0) for item in processed_items) / max(len(processed_items), 1),
        "tax_amount": round(tax_amount, 2),
        "discount": discount,
        "discount_amount": round(discount_amount, 2),
        "total": round(total, 2),
        "status": "draft",
        "due_date": data.get("due_date"),
        "notes": data.get("notes"),
        # Payment verification fields
        "bank": data.get("bank"),
        "expected_account": data.get("expected_account") or data.get("expectedAccount"),
        "currency": data.get("currency", "KHR"),
        "verification_status": "pending",
        "verified_at": None,
        "verified_by": None,
        "verification_note": None,
        "created_at": _now_iso(),
        "updated_at": _now_iso()
    }

    _invoices[tenant_str][invoice_id] = invoice

    # Return with customer info (from same tenant)
    invoice = invoice.copy()
    invoice["tenant_id"] = tenant_str
    tenant_customers = _customers[tenant_str]
    invoice["customer"] = tenant_customers.get(invoice["customer_id"])
    return invoice


def update_invoice(invoice_id: str, data: Dict) -> Optional[Dict]:
    """Update an existing invoice with recalculated totals if items changed."""
    if invoice_id not in _invoices:
        return None

    invoice = _invoices[invoice_id]

    # Update fields
    for key, value in data.items():
        if value is not None and key != "items":
            invoice[key] = value

    # Recalculate totals if items changed
    if "items" in data and data["items"] is not None:
        items = data["items"]
        processed_items = []

        for item in items:
            quantity = float(item.get("quantity", 1))
            unit_price = float(item.get("unit_price", 0))
            tax_rate = float(item.get("tax_rate", 0))
            item_total = quantity * unit_price

            processed_items.append({
                "id": item.get("id", _generate_id()),
                "description": item.get("description", ""),
                "quantity": quantity,
                "unit_price": unit_price,
                "tax_rate": tax_rate,
                "total": round(item_total, 2)
            })

        invoice["items"] = processed_items

        subtotal = sum(item["total"] for item in processed_items)
        tax_amount = sum(
            item["total"] * (item.get("tax_rate", 0) / 100)
            for item in processed_items
        )
        discount = float(invoice.get("discount", 0))
        discount_amount = subtotal * (discount / 100) if discount > 0 else 0

        invoice["subtotal"] = round(subtotal, 2)
        invoice["tax_amount"] = round(tax_amount, 2)
        invoice["discount_amount"] = round(discount_amount, 2)
        invoice["total"] = round(subtotal + tax_amount - discount_amount, 2)

    invoice["updated_at"] = _now_iso()

    # Return with customer info
    result = invoice.copy()
    result["customer"] = _customers.get(invoice["customer_id"])
    return result


def delete_invoice(invoice_id: str) -> bool:
    """Delete an invoice."""
    if invoice_id in _invoices:
        del _invoices[invoice_id]
        return True
    return False


def verify_invoice(invoice_id: str, data: Dict) -> Optional[Dict]:
    """
    Update verification status of an invoice.

    Args:
        invoice_id: The invoice ID
        data: Dict with verificationStatus, verifiedBy, verificationNote

    Returns:
        Updated invoice or None if not found
    """
    if invoice_id not in _invoices:
        return None

    invoice = _invoices[invoice_id]

    # Update verification fields
    verification_status = data.get("verificationStatus") or data.get("verification_status")
    if verification_status:
        invoice["verification_status"] = verification_status

    verified_by = data.get("verifiedBy") or data.get("verified_by")
    if verified_by:
        invoice["verified_by"] = verified_by

    verification_note = data.get("verificationNote") or data.get("verification_note")
    if verification_note is not None:
        invoice["verification_note"] = verification_note

    # Set verified_at timestamp if status is verified
    if verification_status == "verified":
        invoice["verified_at"] = _now_iso()
        # Also update invoice status to paid when verified
        invoice["status"] = "paid"

    invoice["updated_at"] = _now_iso()

    # Return with customer info
    result = invoice.copy()
    result["customer"] = _customers.get(invoice["customer_id"])
    return result


# ============================================================================
# Statistics
# ============================================================================

def get_stats() -> Dict:
    """Get invoice statistics."""
    invoices = list(_invoices.values())

    total_revenue = sum(i["total"] for i in invoices if i["status"] == "paid")
    pending_amount = sum(i["total"] for i in invoices if i["status"] in ["pending", "draft"])
    paid_amount = total_revenue
    overdue_count = len([i for i in invoices if i["status"] == "overdue"])

    return {
        "total_invoices": len(invoices),
        "total_revenue": round(total_revenue, 2),
        "pending_amount": round(pending_amount, 2),
        "paid_amount": round(paid_amount, 2),
        "overdue_count": overdue_count,
        "customers_count": len(_customers)
    }


# ============================================================================
# PDF Generation (Real - using fpdf2)
# ============================================================================

def generate_pdf(invoice_id: str, tenant_id: str) -> Optional[bytes]:
    """
    Generate a professional invoice PDF matching the frontend design.

    Uses fpdf2 library for real PDF generation with proper styling.

    Args:
        invoice_id: The invoice ID to generate PDF for
        tenant_id: The tenant ID - ensures tenant isolation
    """
    import logging
    logger = logging.getLogger(__name__)

    try:
        from fpdf import FPDF
        logger.info(f"✅ fpdf2 imported successfully for invoice {invoice_id}")
    except ImportError as e:
        # Fallback if fpdf2 not installed
        logger.error(f"❌ fpdf2 not installed (ImportError: {e}), using fallback PDF for invoice {invoice_id}")
        return _generate_pdf_fallback(invoice_id, tenant_id)

    # Wrap entire PDF generation in try-catch to log actual errors
    try:
        invoice = get_invoice(tenant_id, invoice_id)
        if not invoice:
            logger.error(f"Invoice not found for PDF generation: {invoice_id}")
            return None

        # Validate invoice data to prevent fpdf2 errors
        if not invoice.get("invoice_number"):
            logger.error(f"Invoice {invoice_id} missing invoice_number")
            return None

        amount = invoice.get("total", 0)
        if not isinstance(amount, (int, float)) or (isinstance(amount, float) and amount != amount):
            logger.error(f"Invoice {invoice_id} has invalid amount: {amount}")
            return None

        # Colors (matching frontend design)
        BLUE = (74, 144, 226)
        DARK = (31, 41, 55)
        GRAY = (107, 114, 128)
        LIGHT_BG = (249, 250, 251)

        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        # Header section with blue background
        pdf.set_fill_color(*BLUE)
        pdf.rect(0, 0, 210, 45, 'F')

        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 24)
        pdf.set_y(12)
        pdf.cell(0, 10, "INVOICE", ln=True, align="C")

        pdf.set_font("Helvetica", "", 14)
        pdf.cell(0, 8, invoice["invoice_number"], ln=True, align="C")

        # Status badge
        status = invoice.get("status", "pending").upper()
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 8, status, ln=True, align="C")

        # Reset colors
        pdf.set_text_color(*DARK)
        pdf.set_y(55)

        # Customer + Due Date row
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(*GRAY)
        pdf.cell(95, 6, "BILL TO", ln=False)
        pdf.cell(95, 6, "DUE DATE", ln=True)

        pdf.set_text_color(*DARK)
        pdf.set_font("Helvetica", "", 11)
        customer = invoice.get("customer", {}) or {}
        customer_name = customer.get("name", "N/A") or "N/A"
        pdf.cell(95, 6, customer_name, ln=False)
        pdf.cell(95, 6, str(invoice.get("due_date", "N/A") or "N/A"), ln=True)

        if customer.get("email"):
            pdf.set_text_color(*GRAY)
            pdf.cell(95, 5, customer["email"], ln=True)
        if customer.get("phone"):
            pdf.cell(95, 5, customer["phone"], ln=True)

        pdf.ln(10)

        # Line items table
        pdf.set_fill_color(*LIGHT_BG)
        pdf.set_text_color(*GRAY)
        pdf.set_font("Helvetica", "B", 9)

        # Table header
        pdf.cell(80, 8, "DESCRIPTION", border="B", fill=True)
        pdf.cell(20, 8, "QTY", border="B", align="C", fill=True)
        pdf.cell(35, 8, "UNIT PRICE", border="B", align="R", fill=True)
        pdf.cell(40, 8, "TOTAL", border="B", align="R", fill=True, ln=True)

        # Table rows
        pdf.set_text_color(*DARK)
        pdf.set_font("Helvetica", "", 10)
        currency = invoice.get("currency", "KHR")

        items = invoice.get("items", [])
        for item in items:
            description = str(item.get("description", ""))[:35]
            quantity = item.get("quantity", 1)
            price = float(item.get("unit_price", 0))
            item_total = float(item.get("total", price * quantity))

            pdf.cell(80, 7, description)
            pdf.cell(20, 7, str(int(quantity)), align="C")
            if currency == "KHR":
                pdf.cell(35, 7, f"{price:,.0f}", align="R")
                pdf.cell(40, 7, f"{item_total:,.0f}", align="R", ln=True)
            else:
                pdf.cell(35, 7, f"${price:.2f}", align="R")
                pdf.cell(40, 7, f"${item_total:.2f}", align="R", ln=True)

        pdf.ln(5)

        # Totals section (right-aligned)
        pdf.set_x(110)
        pdf.set_fill_color(*LIGHT_BG)

        # Subtotal
        subtotal = float(invoice.get("subtotal", invoice.get("total", 0)))
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(45, 7, "Subtotal:", align="R")
        if currency == "KHR":
            pdf.cell(40, 7, f"{subtotal:,.0f} KHR", align="R", ln=True)
        else:
            pdf.cell(40, 7, f"${subtotal:.2f}", align="R", ln=True)

        # Tax
        tax = float(invoice.get("tax_amount", 0))
        if tax:
            pdf.set_x(110)
            pdf.cell(45, 7, "Tax:", align="R")
            if currency == "KHR":
                pdf.cell(40, 7, f"{tax:,.0f} KHR", align="R", ln=True)
            else:
                pdf.cell(40, 7, f"${tax:.2f}", align="R", ln=True)

        # Discount
        discount = float(invoice.get("discount_amount", 0))
        if discount:
            pdf.set_x(110)
            pdf.cell(45, 7, "Discount:", align="R")
            if currency == "KHR":
                pdf.cell(40, 7, f"-{discount:,.0f} KHR", align="R", ln=True)
            else:
                pdf.cell(40, 7, f"-${discount:.2f}", align="R", ln=True)

        # Grand Total
        pdf.set_x(110)
        pdf.set_font("Helvetica", "B", 12)
        total = float(invoice.get("total", 0))
        pdf.cell(45, 10, "TOTAL:", border="T", align="R")
        if currency == "KHR":
            pdf.cell(40, 10, f"{total:,.0f} KHR", border="T", align="R", ln=True)
        else:
            pdf.cell(40, 10, f"${total:.2f}", border="T", align="R", ln=True)

        pdf.ln(10)

        # Payment Information section
        bank = invoice.get("bank")
        expected_account = invoice.get("expected_account")
        recipient_name = invoice.get("recipient_name")

        if bank or expected_account or recipient_name:
            pdf.set_fill_color(*LIGHT_BG)
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_text_color(*DARK)
            pdf.cell(0, 8, "Payment Information", fill=True, ln=True)

            pdf.set_font("Helvetica", "", 10)
            if bank:
                pdf.cell(30, 6, "Bank:")
                pdf.cell(0, 6, str(bank), ln=True)
            if expected_account:
                pdf.cell(30, 6, "Account:")
                pdf.cell(0, 6, str(expected_account), ln=True)
            if recipient_name:
                pdf.cell(30, 6, "Recipient:")
                pdf.cell(0, 6, str(recipient_name), ln=True)

        # Notes section
        notes = invoice.get("notes")
        if notes:
            pdf.ln(10)
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(0, 8, "Notes", ln=True)
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(*GRAY)
            pdf.multi_cell(0, 5, str(notes))

        return bytes(pdf.output())

    except Exception as e:
        # Log actual error with full traceback for debugging
        logger.error(
            f"❌ PDF generation failed for invoice {invoice_id}: {type(e).__name__}: {str(e)}",
            exc_info=True  # Include full stack trace
        )
        return None


def _generate_pdf_fallback(invoice_id: str, tenant_id: str) -> Optional[bytes]:
    """Fallback PDF generation if fpdf2 is not available."""
    invoice = get_invoice(tenant_id, invoice_id)
    if not invoice:
        return None

    invoice_num = invoice["invoice_number"]
    customer_name = (invoice.get("customer", {}) or {}).get("name", "Unknown")
    total = invoice["total"]
    currency = invoice.get("currency", "KHR")

    if currency == "KHR":
        amount_str = f"{total:,.0f} KHR"
    else:
        amount_str = f"${total:.2f}"

    content = f"Invoice: {invoice_num} | Customer: {customer_name} | Total: {amount_str}"

    pdf_content = f"""%PDF-1.4
1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj
2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj
3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >> endobj
4 0 obj << /Length 100 >> stream
BT /F1 12 Tf 50 700 Td ({content}) Tj ET
endstream endobj
5 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj
xref 0 6
0000000000 65535 f
0000000009 00000 n
0000000058 00000 n
0000000115 00000 n
0000000230 00000 n
0000000380 00000 n
trailer << /Size 6 /Root 1 0 R >>
startxref 450
%%EOF"""

    return pdf_content.encode('latin-1')


# ============================================================================
# Export (Real - using openpyxl for XLSX)
# ============================================================================

def export_invoices(
    format: str = "csv",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
) -> bytes:
    """
    Export invoices to CSV or Excel format.

    Uses openpyxl for real XLSX generation with proper styling.
    """
    invoices = list(_invoices.values())

    # Filter by date if provided
    if start_date:
        invoices = [i for i in invoices if i.get("created_at", "") >= start_date]
    if end_date:
        invoices = [i for i in invoices if i.get("created_at", "") <= end_date]

    if format == "csv":
        lines = ["Invoice Number,Customer,Status,Subtotal,Tax,Discount,Total,Due Date,Created"]
        for inv in invoices:
            customer = _customers.get(inv["customer_id"], {})
            customer_name = customer.get("name", "Unknown").replace(",", " ")
            lines.append(
                f'{inv["invoice_number"]},'
                f'{customer_name},'
                f'{inv["status"]},'
                f'{inv["subtotal"]},'
                f'{inv["tax_amount"]},'
                f'{inv["discount_amount"]},'
                f'{inv["total"]},'
                f'{inv.get("due_date", "")},'
                f'{inv["created_at"]}'
            )
        return "\n".join(lines).encode("utf-8")
    else:
        # XLSX export using openpyxl
        return _generate_xlsx_export(invoices)


def _generate_xlsx_export(invoices: List[Dict]) -> bytes:
    """Generate real XLSX file using openpyxl."""
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback if openpyxl not installed
        return b"PK\x03\x04XLSX_REQUIRES_OPENPYXL"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # Header styling (matching frontend blue theme)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4A90E2")
    header_align = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        "Invoice Number", "Customer", "Status", "Currency",
        "Subtotal", "Tax", "Discount", "Total", "Due Date", "Created"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    for row_num, inv in enumerate(invoices, 2):
        customer = _customers.get(inv["customer_id"], {})
        customer_name = customer.get("name", "Unknown")

        ws.cell(row=row_num, column=1, value=inv["invoice_number"])
        ws.cell(row=row_num, column=2, value=customer_name)
        ws.cell(row=row_num, column=3, value=inv["status"])
        ws.cell(row=row_num, column=4, value=inv.get("currency", "KHR"))
        ws.cell(row=row_num, column=5, value=inv["subtotal"])
        ws.cell(row=row_num, column=6, value=inv["tax_amount"])
        ws.cell(row=row_num, column=7, value=inv["discount_amount"])
        ws.cell(row=row_num, column=8, value=inv["total"])
        ws.cell(row=row_num, column=9, value=inv.get("due_date", ""))
        ws.cell(row=row_num, column=10, value=inv["created_at"])

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================================
# Sample Data Seeding
# ============================================================================

def seed_sample_data():
    """
    Populate with sample data for testing.

    Only seeds if the data stores are empty.
    """
    global _invoice_counter

    # Only seed if empty
    if _customers or _invoices:
        return {"seeded": False, "message": "Data already exists"}

    # Sample customers
    sample_customers = [
        {
            "name": "Acme Corporation",
            "email": "billing@acme.com",
            "phone": "+1-555-0100",
            "company": "Acme Corp",
            "address": "123 Business St, New York, NY 10001"
        },
        {
            "name": "TechStart Inc",
            "email": "invoices@techstart.io",
            "phone": "+1-555-0101",
            "company": "TechStart",
            "address": "456 Innovation Blvd, San Francisco, CA 94105"
        },
        {
            "name": "Global Services LLC",
            "email": "accounts@globalservices.com",
            "phone": "+1-555-0102",
            "address": "789 Commerce Ave, Chicago, IL 60601"
        },
        {
            "name": "Local Business Shop",
            "email": "owner@localbiz.com",
            "phone": "+1-555-0103",
            "address": "321 Main St, Austin, TX 78701"
        },
        {
            "name": "Enterprise Solutions Ltd",
            "email": "procurement@enterprise.com",
            "company": "Enterprise Solutions Ltd",
            "address": "555 Corporate Dr, Seattle, WA 98101"
        },
    ]

    created_customers = []
    for cust_data in sample_customers:
        customer = create_customer(cust_data)
        created_customers.append(customer)

    # Sample invoices with varied statuses
    statuses = ["draft", "pending", "paid", "paid", "overdue"]

    for i, customer in enumerate(created_customers):
        invoice_data = {
            "customer_id": customer["id"],
            "items": [
                {
                    "description": "Consulting Services",
                    "quantity": 10 + i * 2,
                    "unit_price": 150.00,
                    "tax_rate": 10
                },
                {
                    "description": "Software License",
                    "quantity": 1,
                    "unit_price": 500.00 + i * 100,
                    "tax_rate": 10
                },
            ],
            "due_date": (datetime.utcnow() + timedelta(days=30 - i * 7)).strftime("%Y-%m-%d"),
            "notes": f"Sample invoice #{i + 1} for testing purposes."
        }

        invoice = create_invoice(invoice_data)

        # Update status
        update_invoice(invoice["id"], {"status": statuses[i % len(statuses)]})

    return {
        "seeded": True,
        "customers": len(created_customers),
        "invoices": len(_invoices)
    }


def clear_all_data():
    """Clear all mock data (for testing reset) - all tenants."""
    global _customers, _invoices, _invoice_counters
    _customers = {}
    _invoices = {}
    _invoice_counters = {}
    return {"cleared": True, "message": "All tenant mock data cleared"}


def get_data_counts() -> Dict:
    """Get current data counts."""
    return {
        "customers": len(_customers),
        "invoices": len(_invoices)
    }